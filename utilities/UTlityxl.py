import openpyxl

def getRowCount(file,sheetName):
    wrkbook=openpyxl.load_workbook(file)
    sheet=wrkbook[sheetName]
    return sheet.max_row

def getColumnCount(file,sheetName):
    wrkbook = openpyxl.load_workbook(file)
    sheet = wrkbook[sheetName]
    return sheet.max_column

def readData(file,sheetName,rownum,columnno):
    wrkbook = openpyxl.load_workbook(file)
    sheet = wrkbook[sheetName]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    wrkbook = openpyxl.load_workbook(file)
    sheet = wrkbook[sheetName]
    sheet.cell(row=rownum,column=columnno).value=data
    wrkbook.save(file)



